from io import BytesIO
from tokenize import Double
from fastapi import FastAPI, Request, Form, HTTPException
from fastapi.templating import Jinja2Templates
from fastapi.staticfiles import StaticFiles
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
import openpyxl
from pydantic import BaseModel
from uuid import UUID, uuid4
from datetime import datetime, time, timedelta
from typing import Optional
import sqlite3
import pytz
import logging

import urllib

from schemas import WorkDay

logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger(__name__)

app = FastAPI()
app.mount("/static", StaticFiles(directory="static"), name="static")
templates = Jinja2Templates(directory="templates")

def format_time_arabic(t: time) -> str:
    time_str = t.strftime('%I:%M %p').lstrip('0')
    return time_str.replace("AM", "ص").replace("PM", "م")

templates.env.filters['time_ar'] = format_time_arabic

DB_PATH = "db.sqlite3"

def init_db():
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS workdays (
            id TEXT PRIMARY KEY,
            date TEXT,
            start_time TEXT,
            end_time TEXT,
            break_hours REAL,
            work_hours REAL,
            driver_name TEXT,
            notes TEXT
        )
    ''')
    conn.commit()
    conn.close()

@app.on_event("startup")
async def startup_event():
    init_db()

@app.get("/", response_class=HTMLResponse)
async def home(request: Request,
               start_date: Optional[str] = None,
               end_date: Optional[str] = None,
               driver_name: Optional[str] = None):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    query = "SELECT * FROM workdays WHERE 1=1"
    params = []
    if start_date:
        query += " AND date >= ?"
        params.append(start_date)
    if end_date:
        query += " AND date <= ?"
        params.append(end_date)
    if driver_name:
        query += " AND driver_name LIKE ?"
        params.append(f"%{driver_name}%")
    query += " ORDER BY date"
    cursor.execute(query, params)
    rows = cursor.fetchall()
    workdays = [WorkDay.from_db_row(row) for row in rows]
    conn.close()
    return templates.TemplateResponse(
        request=request,
        name="home.html",
        context={
            "workdays": workdays,
            "start_date": start_date,
            "end_date": end_date,
            "driver_name": driver_name,
            "sum_work_hours": sum([wd.work_hours.total_seconds() / 3600 for wd in workdays])
        }
    )

@app.post("/create", response_class=RedirectResponse)
async def create_workday(
    request: Request,
    date: str = Form(...),
    start_time: str = Form(...),
    end_time: str = Form(...),
    break_hours: float = Form(...),
    driver_name: str = Form(...),
    notes: str = Form(...)
):
    try:
        date_obj = datetime.strptime(date, "%Y-%m-%d").replace(tzinfo=pytz.UTC)
        start_time_obj = datetime.strptime(start_time, "%H:%M").time()
        end_time_obj = datetime.strptime(end_time, "%H:%M").time()
        break_delta = timedelta(hours=break_hours)
        start_dt = datetime.combine(date_obj.date(), start_time_obj)
        end_dt = datetime.combine(date_obj.date(), end_time_obj)
        if end_dt < start_dt:
            end_dt += timedelta(days=1)
        total_duration = end_dt - start_dt
        work_delta = total_duration - break_delta
        if work_delta.total_seconds() < 0:
            raise ValueError("Work hours cannot be negative.")
        workday = WorkDay(
            id=uuid4(),
            date=date_obj,
            start_time=start_time_obj,
            end_time=end_time_obj,
            break_hours=break_delta,
            work_hours=work_delta,
            driver_name=driver_name,
            notes=notes
        )
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute('''
            INSERT INTO workdays (id, date, start_time, end_time, break_hours, work_hours, driver_name, notes)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            str(workday.id),
            workday.date.isoformat(),
            workday.start_time.strftime("%H:%M:%S"),
            workday.end_time.strftime("%H:%M:%S"),
            workday.break_hours.total_seconds() / 3600,
            workday.work_hours.total_seconds() / 3600,
            workday.driver_name,
            workday.notes
        ))
        conn.commit()
        conn.close()
        return RedirectResponse(url="/", status_code=303)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.post("/delete/{id}", response_class=RedirectResponse)
async def delete_workday(id: UUID):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("DELETE FROM workdays WHERE id = ?", (str(id),))
    conn.commit()
    conn.close()
    return RedirectResponse(url="/", status_code=303)

@app.post("/update/{id}", response_class=RedirectResponse)
async def update_workday(
    id: UUID,
    request: Request,
    date: str = Form(...),
    start_time: str = Form(...),
    end_time: str = Form(...),
    break_hours: float = Form(...),
    driver_name: str = Form(...),
    notes: str = Form(...)
):
    try:
        date_obj = datetime.strptime(date, "%Y-%m-%d").replace(tzinfo=pytz.UTC)
        start_time_obj = datetime.strptime(start_time, "%H:%M").time()
        end_time_obj = datetime.strptime(end_time, "%H:%M").time()
        break_delta = timedelta(hours=break_hours)
        start_dt = datetime.combine(date_obj.date(), start_time_obj)
        end_dt = datetime.combine(date_obj.date(), end_time_obj)
        if end_dt < start_dt:
            end_dt += timedelta(days=1)
        total_duration = end_dt - start_dt
        work_delta = total_duration - break_delta
        if work_delta.total_seconds() < 0:
            raise ValueError("Work hours cannot be negative.")
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute('''
            UPDATE workdays SET
                date = ?,
                start_time = ?,
                end_time = ?,
                break_hours = ?,
                work_hours = ?,
                driver_name = ?,
                notes = ?
            WHERE id = ?
        ''', (
            date_obj.isoformat(),
            start_time_obj.strftime("%H:%M:%S"),
            end_time_obj.strftime("%H:%M:%S"),
            break_delta.total_seconds() / 3600,
            work_delta.total_seconds() / 3600,
            driver_name,
            notes,
            str(id)
        ))
        conn.commit()
        conn.close()
        return RedirectResponse(url="/", status_code=303)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.post("/export_excel", response_class=StreamingResponse)
async def export_excel(
    request: Request,
    title: str = Form(...),
    start_date: Optional[str] = Form(None),
    end_date: Optional[str] = Form(None),
    driver_name: Optional[str] = Form(None)
):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    query = "SELECT * FROM workdays WHERE 1=1"
    params = []
    if start_date:
        query += " AND date >= ?"
        params.append(start_date)
    if end_date:
        query += " AND date <= ?"
        params.append(end_date)
    if driver_name:
        query += " AND driver_name LIKE ?"
        params.append(f"%{driver_name}%")
    cursor.execute(query, params)
    rows = cursor.fetchall()
    workdays = [WorkDay.from_db_row(row) for row in rows]
    conn.close()

    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "بيان عدد ساعات العمل بشركة مواد لتدوير المخلفات بمصنع العز"

    # Add title
    ws['A1'] = ws.title + " - " + title
    ws['A1'].font = openpyxl.styles.Font(size=14, bold=True)
    ws.merge_cells('A1:H1')

    # Add headers
    headers = ["التاريخ", "اليوم", "وقت البداية", "وقت النهاية", "ساعات الاستراحة", "ساعات العمل", "اسم السائق", "ملاحظات"]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=3, column=col, value=header).font = openpyxl.styles.Font(bold=True)

    # Add data
    for row_idx, workday in enumerate(workdays, start=4):
        ws.cell(row=row_idx, column=1, value=workday.date.strftime('%Y-%m-%d'))
        ws.cell(row=row_idx, column=2, value=workday.weekday)
        ws.cell(row=row_idx, column=3, value=format_time_arabic(workday.start_time))
        ws.cell(row=row_idx, column=4, value=format_time_arabic(workday.end_time))
        ws.cell(row=row_idx, column=5, value=workday.break_hours.total_seconds() / 3600)
        ws.cell(row=row_idx, column=6, value=workday.work_hours.total_seconds() / 3600)
        ws.cell(row=row_idx, column=7, value=workday.driver_name)
        ws.cell(row=row_idx, column=8, value=workday.notes)

    # Auto-adjust column widths, skipping merged cells
    for col_num in range(1, len(headers) + 1):  # Columns A to H (8 columns)
        max_length = 0
        column_letter = openpyxl.utils.get_column_letter(col_num)
        for row in range(3, ws.max_row + 1):  # Start from headers (row 3)
            cell = ws[f"{column_letter}{row}"]
            if not isinstance(cell, openpyxl.cell.cell.MergedCell):  # Skip merged cells
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        adjusted_width = min(max_length + 2, 50)  # Cap width at 50 for readability
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    # Generate and encode filename
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    raw_filename = f"{ws.title} {title}.xlsx"
    encoded_filename = urllib.parse.quote(raw_filename)  # URL-encode the filename
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
    )
