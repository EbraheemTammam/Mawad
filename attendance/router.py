from datetime import datetime, time, timedelta
from io import BytesIO
from typing import Optional
from uuid import UUID, uuid4
from fastapi import APIRouter, Form, HTTPException, Request
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from bidi.algorithm import get_display
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

import openpyxl
import pytz
import urllib
import arabic_reshaper

from database import Db
from .schemas import WorkDay

pdfmetrics.registerFont(TTFont('AmiriRegular', 'static/fonts/Amiri-Regular.ttf'))

attendance_router = APIRouter()
templates = Jinja2Templates(directory="templates")
def format_time_arabic(t: time) -> str:
    time_str = t.strftime('%I:%M %p').lstrip('0')
    return time_str.replace("AM", "ص").replace("PM", "م")
templates.env.filters['time_ar'] = format_time_arabic

@attendance_router.get("/", response_class=HTMLResponse)
async def home(
        request: Request,
        start_date: Optional[str] = None,
        end_date: Optional[str] = None,
        driver_name: Optional[str] = None
    ):
    query = "SELECT * FROM workdays WHERE 1=1"
    params = []
    if start_date:
        query += " AND date >= ?"
        params.append(start_date)
    if end_date:
        query += " AND date <= ?"
        # Adjust end_date to include the full day by adding 1 day and using <
        end_date_plus_one = (datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1)).strftime("%Y-%m-%d")
        params.append(end_date_plus_one)
    if driver_name:
        query += " AND driver_name LIKE ?"
        params.append(f"%{driver_name}%")
    query += " ORDER BY date"
    rows = Db.execute_query(query, params)
    workdays = [WorkDay.from_db_row(row) for row in rows]
    return templates.TemplateResponse(
        request=request,
        name="home.html",
        context={
            "workdays": workdays,
            "start_date": start_date,
            "end_date": end_date,
            "driver_name": driver_name,
            "sum_work_hours": sum([wd.work_hours.total_seconds() / 3600 for wd in workdays])
        }
    )

@attendance_router.post("/create", response_class=RedirectResponse)
async def create_workday(
    request: Request,
    date: str = Form(...),
    start_time: str = Form(...),
    end_time: str = Form(...),
    break_hours: float = Form(...),
    driver_name: str = Form(...),
    notes: str = Form(...)
):
    try:
        date_obj = datetime.strptime(date, "%Y-%m-%d").replace(tzinfo=pytz.UTC)
        start_time_obj = datetime.strptime(start_time, "%H:%M").time()
        end_time_obj = datetime.strptime(end_time, "%H:%M").time()
        break_delta = timedelta(hours=break_hours)
        start_dt = datetime.combine(date_obj.date(), start_time_obj)
        end_dt = datetime.combine(date_obj.date(), end_time_obj)
        if end_dt < start_dt:
            end_dt += timedelta(days=1)
        total_duration = end_dt - start_dt
        work_delta = total_duration - break_delta
        if work_delta.total_seconds() < 0:
            raise ValueError("Work hours cannot be negative.")
        workday = WorkDay(
            id=uuid4(),
            date=date_obj,
            start_time=start_time_obj,
            end_time=end_time_obj,
            break_hours=break_delta,
            work_hours=work_delta,
            driver_name=driver_name,
            notes=notes
        )
        Db.execute_command('''
            INSERT INTO workdays (id, date, start_time, end_time, break_hours, work_hours, driver_name, notes)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            str(workday.id),
            workday.date.isoformat(),
            workday.start_time.strftime("%H:%M:%S"),
            workday.end_time.strftime("%H:%M:%S"),
            workday.break_hours.total_seconds() / 3600,
            workday.work_hours.total_seconds() / 3600,
            workday.driver_name,
            workday.notes
        ))
        return RedirectResponse(url="/", status_code=303)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

@attendance_router.post("/delete/{id}", response_class=RedirectResponse)
async def delete_workday(id: UUID):
    Db.execute_command("DELETE FROM workdays WHERE id = ?", (str(id),))
    return RedirectResponse(url="/", status_code=303)

@attendance_router.post("/update/{id}", response_class=RedirectResponse)
async def update_workday(
    id: UUID,
    request: Request,
    date: str = Form(...),
    start_time: str = Form(...),
    end_time: str = Form(...),
    break_hours: float = Form(...),
    driver_name: str = Form(...),
    notes: str = Form(...)
):
    try:
        date_obj = datetime.strptime(date, "%Y-%m-%d").replace(tzinfo=pytz.UTC)
        start_time_obj = datetime.strptime(start_time, "%H:%M").time()
        end_time_obj = datetime.strptime(end_time, "%H:%M").time()
        break_delta = timedelta(hours=break_hours)
        start_dt = datetime.combine(date_obj.date(), start_time_obj)
        end_dt = datetime.combine(date_obj.date(), end_time_obj)
        if end_dt < start_dt:
            end_dt += timedelta(days=1)
        total_duration = end_dt - start_dt
        work_delta = total_duration - break_delta
        if work_delta.total_seconds() < 0:
            raise ValueError("Work hours cannot be negative.")
        Db.execute_command('''
            UPDATE workdays SET
                date = ?,
                start_time = ?,
                end_time = ?,
                break_hours = ?,
                work_hours = ?,
                driver_name = ?,
                notes = ?
            WHERE id = ?
        ''', (
            date_obj.isoformat(),
            start_time_obj.strftime("%H:%M:%S"),
            end_time_obj.strftime("%H:%M:%S"),
            break_delta.total_seconds() / 3600,
            work_delta.total_seconds() / 3600,
            driver_name,
            notes,
            str(id)
        ))
        return RedirectResponse(url="/", status_code=303)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

@attendance_router.post("/export_excel", response_class=StreamingResponse)
async def export_excel(
    request: Request,
    title: str = Form(...),
    start_date: Optional[str] = Form(None),
    end_date: Optional[str] = Form(None),
    driver_name: Optional[str] = Form(None)
):
    query = "SELECT * FROM workdays WHERE 1=1"
    params = []
    if start_date:
        query += " AND date >= ?"
        params.append(start_date)
    if end_date:
        query += " AND date <= ?"
        params.append(end_date)
    if driver_name:
        query += " AND driver_name LIKE ?"
        params.append(f"%{driver_name}%")
    query += " ORDER BY date"  # Assuming you still want date sorting
    rows = Db.execute_query(query, params)
    workdays = [WorkDay.from_db_row(row) for row in rows]

    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "بيان عدد ساعات عمل لودر أحمد تمام بشركة مواد"

    # Set RTL orientation
    ws.sheet_view.rightToLeft = True

    # Define styles matching frontend
    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color='E5E7EB', end_color='E5E7EB', fill_type='solid')  # bg-gray-200
    white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')    # bg-white
    center_align = Alignment(horizontal='center', vertical='center')

    # Add title
    ws['A1'] = ws.title + " - " + title
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = center_align
    ws.merge_cells('A1:I1')
    for col in range(1, 10):  # Apply border to merged title row
        ws.cell(row=1, column=col).border = border_style

    # Add headers
    headers = ["م", "التاريخ", "اليوم", "وقت البداية", "وقت النهاية", "ساعات الاستراحة", "ساعات العمل", "اسم السائق", "ملاحظات"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = border_style
        cell.alignment = center_align

    # Add data
    i = 1
    for row_idx, workday in enumerate(workdays, start=3):
        cells = [
            (1, i),
            (2, workday.date.strftime('%Y-%m-%d')),
            (3, workday.weekday),
            (4, format_time_arabic(workday.start_time)),
            (5, format_time_arabic(workday.end_time)),
            (6, round(workday.break_hours.total_seconds() / 3600, 2)),
            (7, round(workday.work_hours.total_seconds() / 3600, 2)),
            (8, workday.driver_name),
            (9, workday.notes)
        ]
        for col_idx, value in cells:
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = white_fill
            cell.border = border_style
            cell.alignment = center_align
        i = i + 1

    # Auto-adjust column widths, skipping merged cells
    for col_num in range(1, len(headers) + 1):  # Columns A to I (9 columns)
        max_length = 0
        column_letter = openpyxl.utils.get_column_letter(col_num)
        for row in range(3, ws.max_row + 1):  # Start from headers (row 3)
            cell = ws[f"{column_letter}{row}"]
            if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        adjusted_width = min(max_length + 2, 50)  # Cap width at 50
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    raw_filename = f"{ws.title} - {title}.xlsx"
    encoded_filename = urllib.parse.quote(raw_filename)
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",  # Fixed 'routerlication' typo
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
    )


@attendance_router.post("/export_pdf", response_class=StreamingResponse)
async def export_pdf(
    request: Request,
    title: str = Form(...),
    start_date: Optional[str] = Form(None),
    end_date: Optional[str] = Form(None),
    driver_name: Optional[str] = Form(None)
):
    query = "SELECT * FROM workdays WHERE 1=1"
    params = []
    if start_date:
        query += " AND date >= ?"
        params.append(start_date)
    if end_date:
        query += " AND date <= ?"
        params.append(end_date)
    if driver_name:
        query += " AND driver_name LIKE ?"
        params.append(f"%{driver_name}%")
    query += " ORDER BY date"
    rows = Db.execute_query(query, params)
    workdays = [WorkDay.from_db_row(row) for row in rows]

    # Create PDF buffer
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)

    # Styles
    styles = getSampleStyleSheet()
    rtl_style = ParagraphStyle(
        'rtl',
        parent=styles['Normal'],
        fontName='AmiriRegular',  # You'll need to register an Arabic-supporting font
        fontSize=12,
        alignment=2,  # 2 is right-aligned for RTL
        leading=14
    )

    title_style = ParagraphStyle(
        'title',
        parent=styles['Title'],
        fontName='AmiriRegular',
        fontSize=16,
        fontWeight='Bold',
        alignment=1,
        leading=14,
        spaceAfter=12
    )
    # Function to reshape Arabic text
    def reshape_arabic(text):
        if not text:
            return text
        reshaped_text = arabic_reshaper.reshape(str(text))
        return get_display(reshaped_text)

    # Title
    elements = []
    doc_title = f"بيان عدد ساعات عمل لودر أحمد تمام بشركة مواد - {title}"
    reshaped_title = reshape_arabic(doc_title)
    elements.append(Paragraph(reshaped_title, title_style))
    elements.append(Paragraph("<br/><br/>", rtl_style))  # Add some spacing

    # Table data
    headers = ["م", "التاريخ", "اليوم", "وقت البداية", "وقت النهاية", "ساعات الاستراحة", "ساعات العمل", "اسم السائق", "ملاحظات"]
    headers = headers[::-1] # Reverse headers for RTL
    # Reshape headers
    reshaped_headers = [reshape_arabic(header) for header in headers]
    data = [reshaped_headers]

    # Add workday data
    for i, workday in enumerate(workdays, 1):
        row = [
            str(i),
            workday.date.strftime('%Y-%m-%d'),
            reshape_arabic(workday.weekday),
            reshape_arabic(format_time_arabic(workday.start_time)),
            reshape_arabic(format_time_arabic(workday.end_time)),
            str(round(workday.break_hours.total_seconds() / 3600, 2)),
            str(round(workday.work_hours.total_seconds() / 3600, 2)),
            reshape_arabic(workday.driver_name),
            reshape_arabic(workday.notes or "")
        ]
        row = row[::-1] # Reverse row for RTL
        data.append(row)

    # Create table
    table = Table(data)
    table.setStyle(TableStyle([
        ('FONT', (0, 0), (-1, -1), 'AmiriRegular'),  # Use Arabic-supporting font
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('RIGHTPADDING', (0, 0), (-1, -1), 10),
        ('LEFTPADDING', (0, 0), (-1, -1), 10),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
    ]))

    elements.append(table)

    # Build PDF
    doc.build(elements)
    buffer.seek(0)

    # Prepare filename
    raw_filename = f"{doc_title}.pdf"
    encoded_filename = urllib.parse.quote(raw_filename)

    # Return streaming response
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
    )
